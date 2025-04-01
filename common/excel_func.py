# -*- coding: utf-8 -*-
import re

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Union


def get_excel_header(sheet: Worksheet, head_idx: int=1):
    header = []
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=head_idx, column=col_idx)
        if not cell.value:
            break
        header.append(cell.value)
    return header


def excel_find_idx(sheet: Worksheet, col_idxs: List[Union[str, int]], values: List[str], start_idx: int=2, ignore_whitespace: bool=False, ignore_case: bool=False):
    col_idxs = [column_index_from_string(idx) if isinstance(idx, str) else idx for idx in col_idxs]
    find_idx = 0
    new_values = []
    for v in values:
        v = str(v).strip()
        if ignore_whitespace:
            v = re.sub(r'\s+', '', v)
        if ignore_case:
            v = v.lower()
        new_values.append(v)
    for row_idx in range(start_idx, sheet.max_row + 1):
        check_values = []
        for col_idx in col_idxs:
            v = sheet.cell(row=row_idx, column=col_idx).value
            if not v:
                v = ''
                check_values.append(v)
                continue
            v = str(v)
            if ignore_whitespace:
                v = re.sub(r'\s+', '', v)
            if ignore_case:
                v = v.lower()
            check_values.append(v)
        if check_values == new_values:
            find_idx = row_idx
            break
    return find_idx


def excel_re_sub(sheet: Worksheet, col_idxs: List[Union[str, int]], pattern: str, replace: str, start_idx: int=2, flags: int=0, tip: bool=False):
    col_idxs = [column_index_from_string(idx) if isinstance(idx, str) else idx for idx in col_idxs]
    replace_info = {}
    for row_idx in range(start_idx, sheet.max_row + 1):
        for col_idx in col_idxs:
            cell = sheet.cell(row=row_idx, column=col_idx)
            if not isinstance(cell.value, str):
                continue
            if not re.search(pattern, cell.value, flags=flags):
                continue
            new_value = re.sub(pattern, replace, cell.value, flags=flags)
            cell.value = new_value
            if tip:
                if col_idx not in replace_info:
                    replace_info[col_idx] = 0
                replace_info[col_idx] += 1
    if tip:
        if any(replace_info.values()):
            print(f'执行替换{len(col_idxs)}列{sheet.max_row-start_idx+1}行，共替换{sum(replace_info.values())}个数据，替换详情：')
            for col_idx, count in replace_info.items():
                print(f'{get_column_letter(col_idx)}列替换了{count}个数据')
        else:
            print(f'执行替换{len(col_idxs)}列{sheet.max_row-start_idx+1}行，未匹配到任何数据')


def excel_fill_nan(sheet: Worksheet, col_idxs: List[Union[str, int]], value, start_idx: int=2):
    col_idxs = [column_index_from_string(idx) if isinstance(idx, str) else idx for idx in col_idxs]
    for row_idx in range(start_idx, sheet.max_row + 1):
        for col_idx in col_idxs:
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is None or cell.value == '':
                cell.value = value


def excel_forward_fill(sheet: Worksheet, col_idxs: List[Union[str, int]], start_idx: int = 2):
    col_idxs = [column_index_from_string(idx) if isinstance(idx, str) else idx for idx in col_idxs]
    for col_idx in col_idxs:
        prev_value = None
        for row_idx in range(start_idx, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None and cell.value != '':
                prev_value = cell.value
            elif prev_value is not None:
                cell.value = prev_value

def excel_backward_fill(sheet: Worksheet, col_idxs: List[Union[str, int]], start_idx: int = 2):
    col_idxs = [column_index_from_string(idx) if isinstance(idx, str) else idx for idx in col_idxs]
    next_value_dict = {}
    for col_idx in col_idxs:
        next_value = None
        for row_idx in range(sheet.max_row, start_idx - 1, -1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None and cell.value != '':
                next_value = cell.value
            elif next_value is not None:
                cell.value = next_value


def excel_merge_identical_cells(sheet: Worksheet, col_idxs: List[Union[str, int]], start_idx: int = 2):
    col_idxs = [column_index_from_string(idx) if isinstance(idx, str) else idx for idx in col_idxs]
    for col_idx in col_idxs:
        merge_start_row = start_idx
        current_value = None
        for row_idx in range(start_idx, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value != current_value:
                if current_value is not None and merge_start_row < row_idx - 1:
                    sheet.merge_cells(start_row=merge_start_row, start_column=col_idx,
                                      end_row=row_idx - 1, end_column=col_idx)
                merge_start_row = row_idx
                current_value = cell.value
        if current_value is not None and merge_start_row < sheet.max_row:
            sheet.merge_cells(start_row=merge_start_row, start_column=col_idx,
                              end_row=sheet.max_row, end_column=col_idx)
